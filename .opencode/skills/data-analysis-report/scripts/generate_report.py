#!/usr/bin/env python3
"""
Data Analysis Report Generator

Generates styled Excel reports with YoY/MoM analysis from multi-period data.
Supports multi-sheet and multi-file Excel sources.
Optional group-by summary sheet (e.g., supervisor summary).
"""

import pandas as pd
import numpy as np
import sys
import os
import yaml
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass


def load_config(config_path: str) -> dict:
    """Load configuration from YAML file."""
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def load_data(config: dict) -> dict:
    """Load data from Excel based on configuration."""
    data = {}
    ds = config['data_source']
    periods = config.get('periods', {})

    if ds['type'] == 'multi_sheet':
        xl = pd.ExcelFile(ds['path'])
        for sheet in ds['sheets']:
            data[str(sheet)] = pd.read_excel(xl, sheet_name=str(sheet))
    elif ds['type'] == 'multi_file':
        for file_path in ds['files']:
            name = Path(file_path).stem
            df = pd.read_excel(file_path)
            if 'current_file' in periods and name == periods['current_file']:
                data[periods['current']] = df
            elif 'previous_file' in periods and name == periods['previous_file']:
                data[periods['previous']] = df
            else:
                data[name] = df

    return data


def standardize_columns(df: pd.DataFrame, period: str, value_columns: list, key_column: str) -> pd.DataFrame:
    """Rename columns with period suffix."""
    new_cols = {key_column: 'key'}
    for vc in value_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        if col_name in df.columns:
            new_cols[col_name] = f"{col_name}_{period}"
    return df.rename(columns=new_cols)


def merge_periods(data: dict, periods: dict, key_column: str, value_columns: list) -> pd.DataFrame:
    """Merge data from multiple periods using outer join."""
    current_df = standardize_columns(data[periods['current']], 'current', value_columns, key_column)
    previous_df = standardize_columns(data[periods['previous']], 'previous', value_columns, key_column)

    merged = current_df.merge(previous_df, on='key', how='outer')
    merged.rename(columns={'key': key_column}, inplace=True)

    return merged


def calc_comparison(current, previous):
    """
    Calculate comparison percentage.

    Returns 'N/A' if current or previous is NaN, or if previous is 0.
    """
    if pd.isna(current) or pd.isna(previous) or previous == 0:
        return 'N/A'
    return (current - previous) / previous


def calculate_analysis(df: pd.DataFrame, value_columns: list, analysis_types: list) -> pd.DataFrame:
    """Calculate YoY/MoM comparisons for each value column."""
    for vc in value_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        current_col = f"{col_name}_current"
        previous_col = f"{col_name}_previous"

        if 'yoy' in analysis_types:
            df[f"{col_name}_yoy"] = df.apply(
                lambda row: calc_comparison(row[current_col], row[previous_col]), axis=1
            )

    return df


def add_totals(df: pd.DataFrame, key_column: str, value_columns: list, analysis_types: list) -> pd.DataFrame:
    """Add summary row at the end with totals and overall comparison."""
    totals = {key_column: '总计'}

    for vc in value_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        current_col = f"{col_name}_current"
        previous_col = f"{col_name}_previous"

        current_total = df[current_col].sum()
        previous_total = df[previous_col].sum()

        totals[current_col] = current_total
        totals[previous_col] = previous_total

        if 'yoy' in analysis_types:
            totals[f"{col_name}_yoy"] = calc_comparison(current_total, previous_total)

    return pd.concat([df, pd.DataFrame([totals])], ignore_index=True)


def filter_group_metrics(value_columns: list, metrics_config) -> list:
    """
    Filter metrics for group summary.

    Args:
        value_columns: List of value column configs
        metrics_config: 'auto' or explicit list of metric names

    Returns:
        List of metric names for group summary
    """
    if metrics_config == 'auto':
        filtered = []
        for vc in value_columns:
            col_name = vc['name'] if isinstance(vc, dict) else vc
            if '总' in col_name:
                filtered.append(col_name)
        return filtered if filtered else [vc['name'] if isinstance(vc, dict) else vc for vc in value_columns]
    elif isinstance(metrics_config, list):
        return metrics_config
    else:
        return [vc['name'] if isinstance(vc, dict) else vc for vc in value_columns]


def generate_group_summary(data: dict, periods: dict, group_config: dict, 
                           value_columns: list, analysis_types: list) -> pd.DataFrame:
    """
    Generate group-by summary (e.g., supervisor summary).

    Args:
        data: Dict of period DataFrames
        periods: Period labels
        group_config: group_by configuration
        value_columns: List of value column configs
        analysis_types: List of analysis types

    Returns:
        DataFrame with group summary
    """
    group_col = group_config['column']
    null_handling = group_config.get('null_handling', 'ignore')
    metrics = filter_group_metrics(value_columns, group_config.get('metrics', 'auto'))

    current_df = data[periods['current']].copy()
    previous_df = data[periods['previous']].copy()

    if null_handling == 'ignore':
        current_df = current_df[current_df[group_col].notna()]
        previous_df = previous_df[previous_df[group_col].notna()]

    group_current = current_df.groupby(group_col)[metrics].sum().reset_index()
    group_previous = previous_df.groupby(group_col)[metrics].sum().reset_index()

    group_current.rename(columns={m: f"{m}_current" for m in metrics}, inplace=True)
    group_previous.rename(columns={m: f"{m}_previous" for m in metrics}, inplace=True)

    merged = group_current.merge(group_previous, on=group_col, how='outer')

    for m in metrics:
        merged[f"{m}_yoy"] = merged.apply(
            lambda row: calc_comparison(row[f"{m}_current"], row[f"{m}_previous"]), axis=1
        )

    output_cols = [group_col]
    for m in metrics:
        output_cols.extend([f"{m}_previous", f"{m}_current", f"{m}_yoy"])

    result = merged[output_cols]

    rename_map = {group_col: group_col}
    for m in metrics:
        prev_label = periods['previous']
        curr_label = periods['current']
        rename_map[f"{m}_previous"] = f"{prev_label}{m}"
        rename_map[f"{m}_current"] = f"{curr_label}{m}"
        rename_map[f"{m}_yoy"] = f"{m}同比"
    result.rename(columns=rename_map, inplace=True)

    totals = {group_col: '总计'}
    for m in metrics:
        totals[f"{prev_label}{m}"] = result[f"{prev_label}{m}"].sum()
        totals[f"{curr_label}{m}"] = result[f"{curr_label}{m}"].sum()
        totals[f"{m}同比"] = calc_comparison(totals[f"{curr_label}{m}"], totals[f"{prev_label}{m}"])

    result = pd.concat([result, pd.DataFrame([totals])], ignore_index=True)

    return result


def style_workbook(ws, value_columns: list, analysis_types: list):
    """Apply styling to worksheet: header, borders, number formats, total row."""
    header_fill = PatternFill(start_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    total_fill = PatternFill(start_color='FFC000', fill_type='solid')
    total_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    yoy_cols = set()
    int_cols = set()
    col_idx = 2
    for vc in value_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        if '杯数' in col_name:
            int_cols.add(col_idx)
        col_idx += 1
        if '杯数' in col_name:
            int_cols.add(col_idx)
        col_idx += 1
        if 'yoy' in analysis_types:
            yoy_cols.add(col_idx)
            col_idx += 1

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            if col in yoy_cols:
                if cell.value != 'N/A' and cell.value is not None:
                    cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='center')
            elif col in int_cols:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = total_fill
        cell.font = total_font

    ws.column_dimensions['A'].width = 30
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 14


def style_group_sheet(ws, metrics: list):
    """Apply styling to group summary sheet."""
    header_fill = PatternFill(start_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    total_fill = PatternFill(start_color='FFC000', fill_type='solid')
    total_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    yoy_cols = set()
    int_cols = set()
    col_idx = 2
    for m in metrics:
        if '杯数' in m:
            int_cols.add(col_idx)
        col_idx += 1
        if '杯数' in m:
            int_cols.add(col_idx)
        col_idx += 1
        yoy_cols.add(col_idx)
        col_idx += 1

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            if col in yoy_cols:
                if cell.value != 'N/A' and cell.value is not None:
                    cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='center')
            elif col in int_cols:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = total_fill
        cell.font = total_font

    ws.column_dimensions['A'].width = 15
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 14


def generate_report(config_path: str) -> str:
    """
    Main function to generate report from config.

    Args:
        config_path: Path to YAML configuration file

    Returns:
        Path to generated Excel file
    """
    config = load_config(config_path)

    data = load_data(config)
    merged = merge_periods(
        data,
        config['periods'],
        config['key_column'],
        config['value_columns']
    )

    result = calculate_analysis(
        merged,
        config['value_columns'],
        config['analysis']
    )

    result = add_totals(
        result,
        config['key_column'],
        config['value_columns'],
        config['analysis']
    )

    output_cols = [config['key_column']]
    for vc in config['value_columns']:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        output_cols.extend([
            f"{col_name}_previous",
            f"{col_name}_current",
            f"{col_name}_yoy"
        ])

    result = result[output_cols]

    rename_map = {config['key_column']: config['key_column']}
    for vc in config['value_columns']:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        prev_label = config['periods']['previous']
        curr_label = config['periods']['current']
        rename_map[f"{col_name}_previous"] = f"{prev_label}{col_name}"
        rename_map[f"{col_name}_current"] = f"{curr_label}{col_name}"
        rename_map[f"{col_name}_yoy"] = f"{col_name}同比"

    result.rename(columns=rename_map, inplace=True)

    wb = Workbook()
    ws = wb.active
    ws.title = config['output']['title']

    for r in dataframe_to_rows(result, index=False, header=True):
        ws.append(r)

    style_workbook(ws, config['value_columns'], config['analysis'])

    if 'group_by' in config and config['group_by']:
        group_config = config['group_by']
        group_col = group_config['column']
        sheet_name = group_config.get('sheet_name', f"{group_col}汇总")
        metrics = filter_group_metrics(config['value_columns'], group_config.get('metrics', 'auto'))

        group_result = generate_group_summary(
            data,
            config['periods'],
            group_config,
            config['value_columns'],
            config['analysis']
        )

        ws2 = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(group_result, index=False, header=True):
            ws2.append(r)

        style_group_sheet(ws2, metrics)

    date_dir = datetime.now().strftime('%Y%m%d')
    output_dir = Path(config['output']['dir']) / date_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    time_str = datetime.now().strftime('%H%M%S')
    output_file = output_dir / f"{config['output']['title']}_{time_str}.xlsx"
    wb.save(output_file)

    print(f"报表已生成: {output_file}")
    print(f"点位数量: {len(result) - 1}")

    if 'group_by' in config and config['group_by']:
        print(f"分组汇总: {len(group_result) - 1} 条")

    return str(output_file)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python generate_report.py <config.yaml>")
        sys.exit(1)

    generate_report(sys.argv[1])
