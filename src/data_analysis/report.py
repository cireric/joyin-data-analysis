"""
Report Generator Module

Handles generating Excel reports from analyzed data.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List

from .loader import load_data, validate_data
from .analyzer import merge_periods, fill_missing_values, calculate_analysis, add_totals, merge_maintenance_counts, calc_comparison
from .maintenance import process_maintenance_data
from .styler import style_workbook, style_group_sheet, sanitize_filename


def reorder_value_columns(value_columns: list) -> list:
    """
    Reorder columns: priority columns (含"总") first, then others.
    """
    priority = []
    others = []
    
    for vc in value_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        if not col_name:
            continue
        is_priority = '总' in col_name
        
        if is_priority:
            priority.append(vc)
        else:
            others.append(vc)
    
    return priority + others


def filter_group_metrics(value_columns: list, metrics_config) -> list:
    """
    Filter metrics for group summary.

    Args:
        value_columns: List of value column configs
        metrics_config: 'auto' or explicit list of metric names

    Returns:
        List of metric names for group summary
    """
    if metrics_config == 'auto':
        filtered = []
        for vc in value_columns:
            col_name = vc['name'] if isinstance(vc, dict) else vc
            if '总' in col_name:
                filtered.append(col_name)
        return filtered if filtered else [vc['name'] if isinstance(vc, dict) else vc for vc in value_columns]
    elif isinstance(metrics_config, list):
        return metrics_config
    else:
        return [vc['name'] if isinstance(vc, dict) else vc for vc in value_columns]


def generate_group_summary(data: dict, periods: dict, group_config: dict, 
                           value_columns: list, analysis_types: list) -> pd.DataFrame:
    """
    Generate group-by summary (e.g., supervisor summary).

    Args:
        data: Dict of period DataFrames
        periods: Period labels
        group_config: group_by configuration
        value_columns: List of value column configs
        analysis_types: List of analysis types

    Returns:
        DataFrame with group summary
    """
    group_col = group_config['column']
    null_handling = group_config.get('null_handling', 'ignore')
    metrics = filter_group_metrics(value_columns, group_config.get('metrics', 'auto'))

    current_df = data[periods['current']].copy()
    previous_df = data[periods['previous']].copy()

    if group_col not in current_df.columns:
        raise ValueError(f"Group column '{group_col}' not found in current period data")
    if group_col not in previous_df.columns:
        raise ValueError(f"Group column '{group_col}' not found in previous period data")

    if null_handling == 'ignore':
        current_df = current_df[current_df[group_col].notna()]
        previous_df = previous_df[previous_df[group_col].notna()]

    if current_df.empty:
        raise ValueError("Current period data is empty after filtering null group values")
    if previous_df.empty:
        raise ValueError("Previous period data is empty after filtering null group values")

    group_current = current_df.groupby(group_col)[metrics].sum().reset_index()
    group_previous = previous_df.groupby(group_col)[metrics].sum().reset_index()

    group_current.rename(columns={m: f"{m}_current" for m in metrics}, inplace=True)
    group_previous.rename(columns={m: f"{m}_previous" for m in metrics}, inplace=True)

    merged = group_current.merge(group_previous, on=group_col, how='outer')

    for m in metrics:
        merged[f"{m}_yoy"] = merged.apply(
            lambda row: calc_comparison(row[f"{m}_current"], row[f"{m}_previous"]), axis=1
        )

    output_cols = [group_col]
    for m in metrics:
        output_cols.extend([f"{m}_previous", f"{m}_current", f"{m}_yoy"])

    result = merged[output_cols]

    rename_map = {group_col: group_col}
    for m in metrics:
        prev_label = periods['previous']
        curr_label = periods['current']
        rename_map[f"{m}_previous"] = f"{prev_label}{m}"
        rename_map[f"{m}_current"] = f"{curr_label}{m}"
        rename_map[f"{m}_yoy"] = f"{m}同比"
    result.rename(columns=rename_map, inplace=True)

    totals = {group_col: '总计'}
    for m in metrics:
        totals[f"{prev_label}{m}"] = result[f"{prev_label}{m}"].sum()
        totals[f"{curr_label}{m}"] = result[f"{curr_label}{m}"].sum()
        totals[f"{m}同比"] = calc_comparison(totals[f"{curr_label}{m}"], totals[f"{prev_label}{m}"])

    result = pd.concat([result, pd.DataFrame([totals])], ignore_index=True)

    return result


def generate_supervisor_detail(result: pd.DataFrame, group_col: str, 
                               supervisor: str, key_column: str,
                               value_columns: list, analysis_types: list,
                               periods: dict) -> pd.DataFrame:
    """
    Generate detail sheet for a single supervisor.
    
    Args:
        result: Full analysis result DataFrame
        group_col: Group column name (e.g., 督导人员)
        supervisor: Supervisor name to filter
        key_column: Key column name
        value_columns: List of value column configs
        analysis_types: List of analysis types
        periods: Period labels
        
    Returns:
        Filtered DataFrame for the supervisor
    """
    if group_col not in result.columns:
        return None
    
    supervisor_data = result[result[group_col] == supervisor].copy()
    
    if supervisor_data.empty:
        return None
    
    reordered_columns = reorder_value_columns(value_columns)

    output_cols = [key_column]
    for vc in reordered_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        output_cols.extend([
            f"{col_name}_previous",
            f"{col_name}_current",
            f"{col_name}_yoy"
        ])

    supervisor_data = supervisor_data[output_cols]

    point_count = len(supervisor_data)
    totals = {key_column: f'总计（{point_count}台）'}
    for vc in value_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        prev_col = f"{col_name}_previous"
        curr_col = f"{col_name}_current"
        
        prev_total = supervisor_data[prev_col].sum()
        curr_total = supervisor_data[curr_col].sum()
        
        totals[prev_col] = prev_total
        totals[curr_col] = curr_total
        
        if 'yoy' in analysis_types:
            totals[f"{col_name}_yoy"] = calc_comparison(curr_total, prev_total)
    
    supervisor_data = pd.concat([supervisor_data, pd.DataFrame([totals])], ignore_index=True)

    rename_map = {key_column: key_column}
    for vc in value_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        prev_label = periods['previous']
        curr_label = periods['current']
        rename_map[f"{col_name}_previous"] = f"{prev_label}{col_name}"
        rename_map[f"{col_name}_current"] = f"{curr_label}{col_name}"
        rename_map[f"{col_name}_yoy"] = f"{col_name}同比"

    supervisor_data.rename(columns=rename_map, inplace=True)

    return supervisor_data


def generate_report(config: dict, data: dict = None, maintenance_file: str = None) -> tuple:
    """
    Generate Excel report from configuration.
    
    Args:
        config: Configuration dictionary
        data: Pre-loaded data (optional, will load if not provided)
        maintenance_file: Path to maintenance Excel file (optional)
        
    Returns:
        Tuple of (output_file_path, point_count, group_count, supervisor_count, unmatched_count)
    """
    if data is None:
        data = load_data(config)
    
    validate_data(data, config['periods'], config['key_column'], config['value_columns'])
    
    merged = merge_periods(data, config['periods'], config['key_column'], config['value_columns'])
    merged = fill_missing_values(merged, config['key_column'], config['value_columns'])
    
    maintenance_counts = {}
    unmatched_df = None
    unmatched_count = 0
    
    if maintenance_file:
        current_df = data[config['periods']['current']]
        maintenance_config = config.get('maintenance', {})
        code_col = maintenance_config.get('code_column', '机器编号')
        if code_col in current_df.columns:
            maintenance_counts, unmatched_df = process_maintenance_data(
                maintenance_file,
                current_df,
                site_col=maintenance_config.get('site_column', '站点名称'),
                code_col=code_col,
                name_col=config['key_column']
            )
            if unmatched_df is not None:
                unmatched_count = len(unmatched_df)
    
    group_col = None
    if 'group_by' in config and config['group_by']:
        group_col = config['group_by']['column']
        if group_col and group_col not in merged.columns:
            current_df = data[config['periods']['current']]
            if group_col in current_df.columns:
                merged = merged.merge(
                    current_df[[config['key_column'], group_col]],
                    on=config['key_column'],
                    how='left'
                )
    
    result = calculate_analysis(merged, config['value_columns'], config['analysis'])
    
    if maintenance_counts:
        result = merge_maintenance_counts(result, maintenance_counts, config['key_column'])
    
    result = add_totals(result, config['key_column'], config['value_columns'], config['analysis'])
    
    reordered_columns = reorder_value_columns(config['value_columns'])

    output_cols = [config['key_column']]
    if group_col and group_col in result.columns:
        output_cols.append(group_col)
    for vc in reordered_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        output_cols.extend([
            f"{col_name}_previous",
            f"{col_name}_current",
            f"{col_name}_yoy"
        ])
    if maintenance_file:
        output_cols.append('运维次数')

    result_for_output = result[output_cols].copy()
    if group_col and group_col in result_for_output.columns:
        result_for_output = result_for_output.drop(columns=[group_col])

    rename_map = {config['key_column']: config['key_column']}
    for vc in config['value_columns']:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        prev_label = config['periods']['previous']
        curr_label = config['periods']['current']
        rename_map[f"{col_name}_previous"] = f"{prev_label}{col_name}"
        rename_map[f"{col_name}_current"] = f"{curr_label}{col_name}"
        rename_map[f"{col_name}_yoy"] = f"{col_name}同比"

    result_for_output.rename(columns=rename_map, inplace=True)

    wb = Workbook()
    ws = wb.active
    ws.title = config['output']['title']

    for r in dataframe_to_rows(result_for_output, index=False, header=True):
        ws.append(r)

    style_workbook(ws, reordered_columns, config['analysis'])

    group_result = None
    supervisor_count = 0
    if 'group_by' in config and config['group_by']:
        group_config = config['group_by']
        group_col = group_config['column']
        sheet_name = group_config.get('sheet_name', f"{group_col}汇总")
        metrics = filter_group_metrics(config['value_columns'], group_config.get('metrics', 'auto'))

        group_result = generate_group_summary(
            data, config['periods'], group_config, config['value_columns'], config['analysis']
        )

        ws2 = wb.create_sheet(sheet_name)
        for r in dataframe_to_rows(group_result, index=False, header=True):
            ws2.append(r)

        style_group_sheet(ws2, metrics)
        
        supervisors = group_result[group_col].iloc[:-1].tolist()
        supervisors = [s for s in supervisors if pd.notna(s)]
        
        for supervisor in supervisors:
            supervisor_data = generate_supervisor_detail(
                merged, group_col, supervisor, config['key_column'],
                config['value_columns'], config['analysis'], config['periods']
            )
            
            if supervisor_data is not None and not supervisor_data.empty:
                ws_detail = wb.create_sheet(str(supervisor))
                for r in dataframe_to_rows(supervisor_data, index=False, header=True):
                    ws_detail.append(r)
                style_workbook(ws_detail, reordered_columns, config['analysis'])
                supervisor_count += 1

    if unmatched_df is not None and not unmatched_df.empty:
        ws_unmatched = wb.create_sheet('待确认运维')
        for r in dataframe_to_rows(unmatched_df, index=False, header=True):
            ws_unmatched.append(r)

    date_dir = datetime.now().strftime('%Y%m%d')
    output_dir = Path(config['output']['dir']) / date_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    time_str = datetime.now().strftime('%H%M%S')
    safe_title = sanitize_filename(config['output']['title'])
    output_file = output_dir / f"{safe_title}_{time_str}.xlsx"
    wb.save(output_file)

    return str(output_file), len(result) - 1, len(group_result) - 1 if group_result is not None else 0, supervisor_count, unmatched_count