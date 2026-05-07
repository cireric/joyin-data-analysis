"""
Excel Styler Module

Handles styling Excel worksheets with consistent formatting.
"""

import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TOTAL_FILL = PatternFill(start_color='FFC000', fill_type='solid')
TOTAL_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

INTEGER_COLUMN_KEYWORDS = ['杯', '次数', '数量', '台数']


def sanitize_filename(name: str) -> str:
    """
    Sanitize filename by removing dangerous characters.

    Args:
        name: Original filename

    Returns:
        Safe filename string
    """
    sanitized = re.sub(r'[<>:"/\\|?*\.\.]', '_', name)
    sanitized = sanitized.strip()
    return sanitized if sanitized else 'output'


def auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = max(adjusted_width, 8)


def apply_header_style(ws):
    """Apply header styling to first row."""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')


def apply_cell_styles(ws, yoy_cols: set, int_cols: set):
    """Apply cell styling: borders and number formats."""
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER

            if col in yoy_cols:
                if cell.value != 'N/A' and cell.value is not None:
                    cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal='center')
            elif col in int_cols:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')


def apply_total_row_style(ws):
    """Apply total row styling to last row."""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


def style_workbook(ws, value_columns: list, analysis_types: list):
    """
    Apply styling to worksheet: header, borders, number formats, total row.

    Args:
        ws: Worksheet to style
        value_columns: List of value column configs
        analysis_types: List of analysis types (e.g., ['yoy'])
    """
    apply_header_style(ws)

    yoy_cols = set()
    int_cols = set()
    col_idx = 2
    for vc in value_columns:
        col_name = vc['name'] if isinstance(vc, dict) else vc
        is_int = any(kw in col_name for kw in INTEGER_COLUMN_KEYWORDS)
        if is_int:
            int_cols.add(col_idx)
        col_idx += 1
        if is_int:
            int_cols.add(col_idx)
        col_idx += 1
        if 'yoy' in analysis_types:
            yoy_cols.add(col_idx)
            col_idx += 1

    apply_cell_styles(ws, yoy_cols, int_cols)
    apply_total_row_style(ws)
    auto_fit_columns(ws)


def style_group_sheet(ws, metrics: list):
    """
    Apply styling to group summary sheet.

    Args:
        ws: Worksheet to style
        metrics: List of metric names
    """
    apply_header_style(ws)

    yoy_cols = set()
    int_cols = set()
    col_idx = 2
    for m in metrics:
        is_int = any(kw in m for kw in INTEGER_COLUMN_KEYWORDS)
        if is_int:
            int_cols.add(col_idx)
        col_idx += 1
        if is_int:
            int_cols.add(col_idx)
        col_idx += 1
        yoy_cols.add(col_idx)
        col_idx += 1

    apply_cell_styles(ws, yoy_cols, int_cols)
    apply_total_row_style(ws)
    auto_fit_columns(ws)
